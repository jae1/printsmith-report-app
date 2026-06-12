import io
import openpyxl
from openpyxl.styles import Font, PatternFill

def generate_report_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {data['date']}"

    header_font = Font(bold=True, color="FFFFFF")
    sections = [
        ("new_today", "오늘 들어온 잡", "4472C4"),
        ("in_progress", "진행중인 잡", "7030A0"),
        ("ready", "준비된 잡", "70AD47"),
        ("picked_up", "오늘 픽업/배달 완료", "ED7D31"),
        ("paid", "오늘 결제한 잡", "FFC000"),
    ]

    row = 1
    for key, title, color in sections:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Invoice #", "Account", "Contact", "작업 내용"]
        if key == "in_progress":
            headers.append("Status")
        headers.extend(["주문일", "납기일", "금액"])

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor=color)
        row += 1
        
        section_total = 0
        for item in data[key]:
            col = 1
            ws.cell(row=row, column=col, value=item.get("invoicenumber", "")); col += 1
            ws.cell(row=row, column=col, value=item.get("account_display", "-")); col += 1
            ws.cell(row=row, column=col, value=item.get("contact_display", "")); col += 1
            ws.cell(row=row, column=col, value=item.get("job_name", "")); col += 1
            
            if key == "in_progress":
                ws.cell(row=row, column=col, value=item.get("status", "-")); col += 1
            
            ordered = item.get("ordereddate")
            ws.cell(row=row, column=col, value=str(ordered).split("T")[0] if ordered else ""); col += 1
            
            wanted = item.get("wanteddate") or item.get("pickupdate")
            ws.cell(row=row, column=col, value=str(wanted).split("T")[0] if wanted else ""); col += 1
            
            amt = float(item.get("grandtotal", 0) or 0)
            ws.cell(row=row, column=col, value=amt)
            section_total += amt
            row += 1
            
        total_col = 5 if key == "in_progress" else 4
        ws.cell(row=row, column=total_col + 1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=total_col + 2, value=section_total).font = Font(bold=True)
        row += 2

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
